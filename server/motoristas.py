"""
Motoristas — planilha simples (nome, telefone) usada em dois lugares:
  1. Dropdown de seleção de motorista ao criar uma entrega (Plano de
     Entrega), pra gravar motorista/telefoneMotorista na entrega.
  2. Canal WhatsApp (whatsapp.py) — busca reversa por telefone pra
     identificar se quem mandou mensagem é motorista.

Planilha (aba única, cabeçalho na linha 1, dados a partir da linha 2):
  Col A: Nome
  Col B: Telefone
"""
import io
import re
import openpyxl


def _cel_para_texto(v):
    """Converte valor de célula do Excel pra texto de forma segura.
    Célula formatada como número decimal (float) representando um
    inteiro preservaria o '.0' com str(v) direto (ex.: telefone
    5521993822539.0), corrompendo o dado na hora de limpar caracteres
    não-numéricos. Detecta esse caso e converte pra int primeiro."""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _normaliza_telefone(tel):
    return re.sub(r'\D', '', str(tel or ''))


def _parece_telefone(valor):
    """True se o valor, uma vez limpo de pontuação, tem cara de telefone
    (10+ dígitos — DDD+número, com ou sem DDI)."""
    return len(_normaliza_telefone(valor)) >= 10


def ler_motoristas(xlsx_bytes):
    """Devolve lista [{'nome':, 'telefone':}], na ordem da planilha.

    FIX (22/08/2026): a planilha pode vir COM ou SEM linha de cabeçalho —
    detecta automaticamente olhando se a coluna B da primeira linha tem
    cara de telefone (nesse caso, é dado de verdade, não cabeçalho, e a
    leitura começa da linha 1; senão, pula a primeira linha como
    cabeçalho, como antes)."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return []

    primeira = linhas[0]
    tem_cabecalho = not (len(primeira) > 1 and _parece_telefone(primeira[1]))
    dados = linhas[1:] if tem_cabecalho else linhas

    out = []
    for row in dados:
        if not row or not row[0]:
            continue
        nome = _cel_para_texto(row[0])
        telefone = _cel_para_texto(row[1]) if len(row) > 1 else ''
        if nome:
            out.append({'nome': nome, 'telefone': telefone})
    return out


def buscar_motorista_por_telefone(xlsx_bytes, telefone):
    """Confere se um telefone bate com algum motorista cadastrado."""
    alvo = _normaliza_telefone(telefone)
    for m in ler_motoristas(xlsx_bytes):
        if _normaliza_telefone(m['telefone']) == alvo:
            return m
    return None
