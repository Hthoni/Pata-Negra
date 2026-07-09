"""
Tabela mestra de produtos: de-para código interno -> nome MASTER (nome
oficial Pata Negra). Uma planilha (MASTER.xlsx) no bucket de perfis,
editável pelo usuário.

Layout: cabeçalho na linha 1; a partir da linha 2, coluna A = nome master,
colunas B em diante = TODOS os códigos internos que apontam para esse
produto (resolve códigos duplicados do sistema legado).

Usado em duas frentes: troca do nome nos PDF/Excel de embarque e o
somatório por produto no simulador. O mapa é cacheado em memória e
recarregado no upload.
"""

import io
import openpyxl
from storage import master_existe, carregar_master_bytes

_cache = None   # {codigo_str: nome_master}
_ordem = None   # [nome_master, ...] na ordem da planilha
_raizes = None  # {nome_master: raiz}  (raiz explícita da coluna A, layout novo)


def _norm(v):
    """Normaliza um valor de célula de código para string (163.0 -> '163')."""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _eh_codigo(s):
    """True se a string parece um código interno (só dígitos, ex.: '163')."""
    return bool(s) and s.replace('.', '', 1).isdigit()


def _carregar():
    """Lê o MASTER.xlsx. Detecta o layout automaticamente:
      - ANTIGO: A = nome master, B em diante = códigos.
      - NOVO:   A = raiz, B = nome master, C em diante = códigos.
    A detecção olha a 1ª linha de dados: se a coluna B for um código
    (numérico) é o layout antigo; se for texto (o nome do produto), é o novo.
    Assim a mesma função lê os dois formatos e a transição não quebra nada."""
    mapa, ordem, raizes = {}, [], {}
    if not master_existe():
        return mapa, ordem, raizes
    wb = openpyxl.load_workbook(io.BytesIO(carregar_master_bytes()), data_only=True)
    ws = wb['Produtos MASTER'] if 'Produtos MASTER' in wb.sheetnames else wb[wb.sheetnames[0]]

    # detecta o layout na 1ª linha de dados com A e B preenchidos
    layout_novo = False
    for r in range(2, ws.max_row + 1):
        a = _norm(ws.cell(r, 1).value)
        b = _norm(ws.cell(r, 2).value)
        if not a:
            continue
        if b and not _eh_codigo(b):
            layout_novo = True
        break

    col_nome = 2 if layout_novo else 1
    col_cod_ini = 3 if layout_novo else 2
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, col_nome).value
        if not nome or not str(nome).strip():
            continue
        nome = str(nome).strip()
        ordem.append(nome)
        if layout_novo:
            rz = str(ws.cell(r, 1).value or '').strip()
            if rz:
                raizes[nome] = rz
        for c in range(col_cod_ini, ws.max_column + 1):
            s = _norm(ws.cell(r, c).value)
            if _eh_codigo(s):
                mapa[s] = nome
    return mapa, ordem, raizes


def _garantir():
    global _cache, _ordem, _raizes
    if _cache is None:
        _cache, _ordem, _raizes = _carregar()


def get_mapa():
    _garantir()
    return _cache


def get_ordem():
    """Lista de nomes master na ordem da planilha."""
    _garantir()
    return _ordem


def recarregar():
    global _cache, _ordem, _raizes
    _cache, _ordem, _raizes = _carregar()
    return _cache


def nome_master(codigo, fallback=''):
    """Nome master de um código interno; fallback (ex.: nome do cliente) se não mapeado."""
    return get_mapa().get(_norm(codigo), fallback)


# Sufixos de embalagem que definem a variação — removê-los revela a "raiz"
# (o produto físico, somado para planejamento de produção).
_SUFIXOS = ['GRANEL', 'PORCIONADO', 'PORCIONADA', 'DO SEU JEITO 400G',
            'DO SEU JEITO', 'PACT. 400G', '400G', '500G']


def raiz(nome):
    """Raiz (produto base) de um nome master, para o agrupamento/soma no
    Total de Pedidos. Prioridade:
      1) raiz EXPLÍCITA da coluna A do MASTER (o jeito certo, editável pelo
         usuário — cobre variações que o chute não pega, ex.: 'BACON CX 10 KGS');
      2) fallback: tira o sufixo de embalagem de uma lista (compatibilidade
         com o MASTER antigo, sem coluna de raiz)."""
    _garantir()
    n = str(nome or '').strip()
    rz = (_raizes or {}).get(n)
    if rz:
        return rz
    import re
    for s in sorted(_SUFIXOS, key=len, reverse=True):
        r = re.sub(r'\s+' + re.escape(s) + r'\s*$', '', n)
        if r != n:
            return r.strip()
    return n
