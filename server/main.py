"""
Pata Negra — Sistema de Processamento de Pedidos
# rebuild trigger summer
Servidor Flask: recebe pedidos de clientes em PDF, faz o parsing,
e gera Excel (upload interno) + PDF (expedição) padronizados.

Arquitetura modular:
  storage.py    -> persistência de perfis no Google Cloud Storage
  perfil.py     -> leitura do Perfil Excel + matching de produtos
  excel_gen.py  -> geração do Excel de upload
  pdf_gen.py    -> geração do PDF de expedição
  parsers/      -> um parser por cliente (isolados entre si)
"""
from flask import Flask, request, jsonify, send_file
import openpyxl
import io
import base64
import datetime
import traceback
from flask_cors import CORS

from storage import perfil_existe, salvar_perfil, carregar_perfil_bytes, perfil_filename, salvar_romaneio, listar_romaneios, deletar_romaneio, salvar_pedido_pdf, carregar_pedido_pdf, salvar_pedido_excel, carregar_pedido_excel, salvar_geofence, listar_geofences, deletar_geofence, salvar_master, master_existe, carregar_master_bytes, atualizar_status_romaneio, salvar_entrega, carregar_entrega, listar_entregas, deletar_entrega, registrar_desfecho_entrega, salvar_estoque, carregar_estoque, carregar_usuarios, salvar_usuarios, hash_senha, verifica_senha, registrar_login, criar_sessao, validar_sessao, encerrar_sessao
from perfil import ler_perfil, ler_filiais, buscar_filial, ler_operadores
from excel_gen import gerar_excel
from pdf_gen import gerar_pdf, _kg_pdf, gerar_pdf_totais
import master

import re
import re as _re
def _normaliza_cnpj(cnpj):
    """Remove formatação do CNPJ, deixando só dígitos. Definida aqui
    localmente pra não depender da versão do perfil.py no servidor."""
    return _re.sub(r'\D', '', str(cnpj or ''))
import importlib
import pkgutil
import parsers

app = Flask(__name__)
CORS(app)

# Descoberta automática de parsers: qualquer módulo em parsers/ que
# exponha uma função parse() é registrado automaticamente.
# O nome de exibição pode ser definido como __cliente_nome__ no módulo;
# caso contrário usa o nome do módulo com capitalização automática.
CLIENTES = {}
for _info in pkgutil.iter_modules(parsers.__path__):
    if _info.name.startswith('_'):
        continue
    try:
        _mod = importlib.import_module(f'parsers.{_info.name}')
        if not hasattr(_mod, 'parse'):
            continue
        _nome = getattr(_mod, '__cliente_nome__', _info.name.replace('_', ' ').title())
        CLIENTES[_info.name] = {'nome': _nome, 'parse': _mod.parse}
    except Exception as _e:
        print(f'[WARN] parser {_info.name} não carregado: {_e}')

# Registro de clientes sem PDF (pedido lançado manualmente no popup, ex:
# pedidos por telefone/WhatsApp). Mesma chave usada em /perfil/<cliente> e
# nos endpoints /operadores, /filiais, /produtos e /processar-manual abaixo.
# multiFilial=True: cliente tem várias lojas (tabela M:N:O:P:Q do Perfil),
# o popup pede pra escolher a filial pelo nome.
# multiFilial=False: cliente tem CNPJ/endereço único — vem direto do
# cabeçalho do Perfil (linhas 3-6), sem seleção nenhuma no popup.
CLIENTES_MANUAIS = {
    'guanabara_lojas': {'nome': 'Guanabara Lojas', 'multiFilial': True},
    'mundial_lojas': {'nome': 'Mundial Lojas', 'multiFilial': True},
    'guanabara_central': {'nome': 'Guanabara Central', 'multiFilial': False},
    'mundial_central': {'nome': 'Mundial Central', 'multiFilial': False},
    'prezunic': {'nome': 'Prezunic', 'multiFilial': False},
    'soberano': {'nome': 'Soberano', 'multiFilial': True},
}


def _cliente_meta(cliente):
    """Meta do cliente para o fluxo manual, aceitando tanto clientes só-manuais
    quanto clientes com parser (que agora tambem podem lancar pedido manual).
    Parser assume multiFilial=True (tem varias lojas no Perfil)."""
    if cliente in CLIENTES_MANUAIS:
        return CLIENTES_MANUAIS[cliente]
    if cliente in CLIENTES:
        return {'nome': CLIENTES[cliente]['nome'], 'multiFilial': True}
    return None


def _gerar_arquivos_por_empresa(dados, filiais, logo_bytes=None):
    """Detecta split por empresa (produtos de Indústria e Distribuidora no
    mesmo pedido) e gera os pares Excel+PDF correspondentes. Compartilhado
    entre /processar (pedidos parseados de PDF) e /processar-manual
    (pedidos lançados na tela, sem PDF) — a lógica de split não depende de
    como os itens chegaram, só do campo 'empresa' de cada item."""
    empresas_nos_itens = set(
        it.get('empresa') or dados.get('empresa', 2)
        for f in filiais for it in f['itens']
    )
    empresas_nos_itens.discard(None)
    if not empresas_nos_itens:
        empresas_nos_itens = {dados.get('empresa', 2)}

    arquivos = []
    eb_simples = pb_simples = None  # guarda o único par gerado quando não há split, p/ reaproveitar
    for emp_split in sorted(empresas_nos_itens):
        override = emp_split if len(empresas_nos_itens) > 1 else None
        eb = gerar_excel(dados, empresa_override=override)
        pb = gerar_pdf(dados, empresa_override=override, logo_bytes=logo_bytes)
        if len(empresas_nos_itens) == 1:
            eb_simples, pb_simples = eb, pb
        label = ('Indústria' if emp_split == 1 else 'Distribuidora') if len(empresas_nos_itens) > 1 else ''
        arquivos.append({
            'empresa': emp_split,
            'label': label,
            'excel': base64.b64encode(eb).decode(),
            'pdf': base64.b64encode(pb).decode(),
        })
    return arquivos, eb_simples, pb_simples, len(empresas_nos_itens) > 1


@app.route('/romaneios')
def get_romaneios():
    """Lista romaneios para o mapa/pedidos (sem os itens, p/ payload leve).
    Por padrao devolve os ATIVOS (pendente + em_rota). ?status=entregue,falhou
    devolve o historico."""
    try:
        filtro = request.args.get('status')
        alvos = set(s.strip() for s in filtro.split(',')) if filtro else {'pendente', 'em_rota'}
        roms = listar_romaneios()
        out = []
        for r in roms:
            r.setdefault('status', 'pendente')
            if r.get('status') in alvos:
                r.pop('itens', None)
                out.append(r)
        return jsonify(out)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/simular-totais', methods=['POST'])
def simular_totais():
    """Recebe IDs de romaneios do simulador, soma kg fisico por nome MASTER
    e devolve um PDF (lista de producao consolidada + pedidos incluidos)."""
    body = request.get_json(silent=True) or {}
    ids = body.get('ids', [])
    if not ids:
        return jsonify({'erro': 'Nenhum pedido selecionado'}), 400
    try:
        idx = {r['id']: r for r in listar_romaneios()}
        agregado = {}     # nome_master -> kg
        sem_codigo = {}   # nome_cliente -> kg (itens sem codigo ou nao mapeados)
        pedidos_incl = []
        sem_itens = 0
        for rid in ids:
            r = idx.get(rid)
            if not r:
                continue
            pedidos_incl.append({'cliente': r.get('clienteNome') or r.get('cliente') or '',
                                 'filial': r.get('filial', '')})
            its = r.get('itens')
            if not its:
                sem_itens += 1
                continue
            for it in its:
                cod = str(it.get('cod') or '').strip()
                kg = float(it.get('kg') or 0)
                nome_cli = str(it.get('nome') or '').strip()
                nome = master.nome_master(cod, '') if cod else ''
                if nome:
                    agregado[nome] = agregado.get(nome, 0) + kg
                else:
                    chave = nome_cli or ('cod ' + cod if cod else '(sem nome)')
                    sem_codigo[chave] = sem_codigo.get(chave, 0) + kg
        produtos = [{'nome': n, 'kg': round(k, 1), 'alerta': False} for n, k in agregado.items()]
        produtos += [{'nome': n + '  (SEM CODIGO)', 'kg': round(k, 1), 'alerta': True} for n, k in sem_codigo.items()]
        produtos.sort(key=lambda p: p['nome'].lower())
        meta = {
            'data': datetime.datetime.utcnow().strftime('%d/%m/%Y %H:%M') + ' UTC',
            'nPedidos': len(pedidos_incl),
            'totalKg': round(sum(p['kg'] for p in produtos), 1),
            'semItens': sem_itens,
            'titulo': (body.get('titulo') or '').strip(),  # ex.: 'Relatório de Pedidos em Rota'
        }
        pdf = gerar_pdf_totais(produtos, pedidos_incl, meta)
        return send_file(io.BytesIO(pdf), mimetype='application/pdf',
                         as_attachment=False, download_name='simulacao_totais.pdf')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@app.route('/demandas')
def demandas():
    """Demanda de produção: soma o kg físico por produto (nome master) dos
    pedidos PENDENTES (em rota não entram), na ordem da planilha master.
    Opcional ?clientes=A,B para filtrar. Devolve também a raiz de cada produto
    (para o subtotal por agrupamento) e a lista de clientes disponível."""
    try:
        filtro = request.args.get('clientes')
        alvos = set(s.strip() for s in filtro.split('|') if s.strip()) if filtro else None

        pend = [r for r in listar_romaneios() if (r.get('status') or 'pendente') == 'pendente']
        clientes = sorted(set((r.get('clienteNome') or r.get('cliente') or '—') for r in pend))

        agregado = {}   # nome_master -> kg
        for r in pend:
            cli = r.get('clienteNome') or r.get('cliente') or '—'
            if alvos is not None and cli not in alvos:
                continue
            for it in (r.get('itens') or []):
                cod = str(it.get('cod') or '').strip()
                kg = float(it.get('kg') or 0)
                if not cod:
                    continue
                nome = master.nome_master(cod, '')
                if nome:
                    agregado[nome] = agregado.get(nome, 0) + kg

        ordem = master.get_ordem()
        linhas = [{'nome': n, 'raiz': master.raiz(n), 'vol': round(agregado.get(n, 0), 1)} for n in ordem]
        total = round(sum(l['vol'] for l in linhas), 1)
        return jsonify({'ordem': linhas, 'clientes': clientes, 'totalGeral': total})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@app.route('/estoque', methods=['GET'])
def get_estoque():
    """Estoque atual de produtos acabados (por nome master)."""
    try:
        return jsonify(carregar_estoque())
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/estoque', methods=['POST'])
def set_estoque():
    """Salva o estoque. Body: {itens: {nome_master: kg}}. Carimba data/hora."""
    body = request.get_json(silent=True) or {}
    itens = body.get('itens') or {}
    try:
        limpo = {}
        for nome, kg in itens.items():
            try:
                v = float(kg)
            except (TypeError, ValueError):
                v = 0
            if v:
                limpo[nome] = v
        dados = {'itens': limpo, 'atualizadoEm': datetime.datetime.utcnow().isoformat()}
        salvar_estoque(dados)
        return jsonify({'ok': True, 'atualizadoEm': dados['atualizadoEm']})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@app.route('/embarque')
def embarque():
    """Painel de embarque: por produto (ordem da master, linha a linha),
    o estoque de produtos acabados, o volume a EMBARCAR (pedidos das entregas
    em PLANEJAMENTO — as já despachadas/em rota saem da conta, pois já deixaram
    a fábrica) e o quanto falta produzir (embarcar - estoque, mínimo 0).
    A key da resposta segue 'emRota' por compatibilidade com o front."""
    try:
        # pedidos das entregas ativas em planejamento (em rota já foi embarcado)
        ids_plan = set()
        for e in listar_entregas():
            if e.get('status', 'em_andamento') == 'finalizada':
                continue
            if (e.get('fase') or 'planejamento') != 'planejamento':
                continue
            for pid in e.get('pedidoIds', []):
                ids_plan.add(pid)
        agregado = {}
        for r in listar_romaneios():
            if r.get('id') not in ids_plan:
                continue
            for it in (r.get('itens') or []):
                cod = str(it.get('cod') or '').strip()
                kg = float(it.get('kg') or 0)
                if not cod:
                    continue
                nome = master.nome_master(cod, '')
                if nome:
                    agregado[nome] = agregado.get(nome, 0) + kg

        est = carregar_estoque()
        estoque_itens = est.get('itens', {})
        ordem = master.get_ordem()
        linhas = []
        total_falta = 0
        for n in ordem:
            emr = round(agregado.get(n, 0), 1)
            estq = round(float(estoque_itens.get(n, 0)), 1)
            falta = emr - estq
            if falta < 0:
                falta = 0
            total_falta += falta
            linhas.append({'nome': n, 'estoque': estq, 'emRota': emr, 'falta': round(falta, 1)})
        return jsonify({'linhas': linhas, 'totalFalta': round(total_falta, 1),
                        'atualizadoEm': est.get('atualizadoEm')})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500



@app.route('/romaneio/<romaneio_id>', methods=['DELETE'])
def delete_romaneio(romaneio_id):
    """Marca pedido como entregue deletando o pin do mapa."""
    try:
        ok = deletar_romaneio(romaneio_id)
        if ok:
            return jsonify({'ok': True})
        return jsonify({'erro': 'Romaneio não encontrado'}), 404
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/romaneio/<romaneio_id>/status', methods=['POST'])
def set_status_romaneio(romaneio_id):
    """Atualiza o status do pedido sem apagar: pendente/em_rota/entregue/falhou.
    Substitui o antigo DELETE do botao 'Entregue' (agora arquiva em vez de destruir)."""
    body = request.get_json(silent=True) or {}
    novo = (body.get('status') or '').strip()
    if novo not in ('pendente', 'em_rota', 'entregue', 'falhou'):
        return jsonify({'erro': 'status invalido'}), 400
    try:
        # descobre se o pedido está numa entrega (para registrar o desfecho lá)
        idx = {r['id']: r for r in listar_romaneios()}
        rom = idx.get(romaneio_id)
        entrega_id = rom.get('entregaId') if rom else None

        if novo == 'falhou':
            # falha: pedido volta para PENDENTE (mantém data original), registra a falha.
            ok = atualizar_status_romaneio(romaneio_id, 'pendente', falha=True)
            resultado = 'pendente'
        else:
            ok = atualizar_status_romaneio(romaneio_id, novo)
            resultado = novo
        if not ok:
            return jsonify({'erro': 'Romaneio não encontrado'}), 404

        # se estava numa entrega e o desfecho é definitivo, registra e arquiva se esvaziou
        finalizada = False
        if entrega_id and novo in ('entregue', 'falhou'):
            snap = {'cliente': (rom.get('clienteNome') or rom.get('cliente') or ''),
                    'filial': rom.get('filial', ''), 'kg': rom.get('kgPlanejados', 0)}
            _, finalizada = registrar_desfecho_entrega(entrega_id, romaneio_id, novo, snap)
        return jsonify({'ok': True, 'status': resultado, 'entregaFinalizada': finalizada})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@app.route('/entregas', methods=['GET'])
def get_entregas():
    """Lista entregas. Por padrão as ativas (em_andamento). ?fase=planejamento|em_rota
    filtra as ativas por fase (entrega sem fase conta como 'planejamento').
    ?status=finalizada devolve o histórico (últimos 20 dias)."""
    try:
        filtro = (request.args.get('status') or 'em_andamento').strip()
        fase = (request.args.get('fase') or '').strip()
        todas = listar_entregas()
        if filtro == 'finalizada':
            limite = datetime.datetime.utcnow() - datetime.timedelta(days=20)
            out = []
            for e in todas:
                if e.get('status') != 'finalizada':
                    continue
                try:
                    fim = datetime.datetime.fromisoformat(e.get('finalizadaEm', ''))
                    if fim >= limite:
                        out.append(e)
                except Exception:
                    out.append(e)
            out.sort(key=lambda x: x.get('finalizadaEm', ''), reverse=True)
            return jsonify(out)
        # ativas: em_andamento (default) — trata ausência de status como ativa
        ativas = [e for e in todas if e.get('status', 'em_andamento') != 'finalizada']
        if fase:
            ativas = [e for e in ativas if (e.get('fase') or 'planejamento') == fase]
        return jsonify(ativas)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/entregas', methods=['POST'])
def criar_entrega():
    """Cria uma entrega a partir de uma seleção do simulador.
    Body: {nome, pedidoIds:[...]}. Marca cada pedido como em_rota (com trava:
    pedido que ja esta em outra entrega e recusado)."""
    body = request.get_json(silent=True) or {}
    nome = (body.get('nome') or '').strip()
    ids = body.get('pedidoIds') or []
    if not nome:
        return jsonify({'erro': 'Dê um nome à entrega'}), 400
    if not ids:
        return jsonify({'erro': 'Nenhum pedido selecionado'}), 400
    try:
        # trava: verifica se algum pedido ja esta em rota (em outra entrega)
        idx = {r['id']: r for r in listar_romaneios()}
        conflitos = []
        for pid in ids:
            r = idx.get(pid)
            if r and r.get('status') == 'em_rota':
                conflitos.append({'id': pid, 'entrega': r.get('entregaNome', '?'),
                                  'cliente': r.get('clienteNome') or r.get('cliente'),
                                  'filial': r.get('filial', '')})
        if conflitos:
            return jsonify({'erro': 'Há pedidos já em outra entrega', 'conflitos': conflitos}), 409

        eid = f"ENT-{datetime.datetime.utcnow().strftime('%Y%m%d-%H%M%S')}"
        entrega = {'id': eid, 'nome': nome, 'criadaEm': datetime.datetime.utcnow().isoformat(),
                   'status': 'em_andamento', 'fase': 'planejamento', 'pedidoIds': ids}
        salvar_entrega(eid, entrega)
        # marca os pedidos como em_rota, carimbando a entrega
        for pid in ids:
            r = idx.get(pid)
            if not r:
                continue
            r['status'] = 'em_rota'
            r['entregaId'] = eid
            r['entregaNome'] = nome
            salvar_romaneio(pid, r)
        return jsonify({'ok': True, 'entrega': entrega})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@app.route('/entregas/<entrega_id>', methods=['DELETE'])
def excluir_entrega(entrega_id):
    """Desfaz uma entrega: devolve os pedidos para pendente e apaga a entrega."""
    try:
        ent = carregar_entrega(entrega_id)
        if not ent:
            return jsonify({'erro': 'Entrega não encontrada'}), 404
        idx = {r['id']: r for r in listar_romaneios()}
        for pid in ent.get('pedidoIds', []):
            r = idx.get(pid)
            if r and r.get('status') == 'em_rota':
                r['status'] = 'pendente'
                r.pop('entregaId', None)
                r.pop('entregaNome', None)
                salvar_romaneio(pid, r)
        deletar_entrega(entrega_id)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/entregas/<entrega_id>/adicionar', methods=['POST'])
def adicionar_pedido_entrega(entrega_id):
    """Adiciona UM pedido a uma entrega já salva. Body: {pedidoId}.
    Marca o pedido em_rota carimbando a entrega. Trava: recusa se o pedido
    já está em outra entrega."""
    body = request.get_json(silent=True) or {}
    pid = (body.get('pedidoId') or '').strip()
    if not pid:
        return jsonify({'erro': 'pedidoId ausente'}), 400
    try:
        ent = carregar_entrega(entrega_id)
        if not ent:
            return jsonify({'erro': 'Entrega não encontrada'}), 404
        idx = {r['id']: r for r in listar_romaneios()}
        r = idx.get(pid)
        if not r:
            return jsonify({'erro': 'Pedido não encontrado'}), 404
        if r.get('status') == 'em_rota' and r.get('entregaId') != entrega_id:
            return jsonify({'erro': 'Pedido já está em outra entrega',
                            'entrega': r.get('entregaNome', '?')}), 409
        ids = ent.get('pedidoIds', [])
        if pid not in ids:
            ids.append(pid)
            ent['pedidoIds'] = ids
            salvar_entrega(entrega_id, ent)
        r['status'] = 'em_rota'
        r['entregaId'] = entrega_id
        r['entregaNome'] = ent.get('nome', '')
        salvar_romaneio(pid, r)
        return jsonify({'ok': True, 'entregaNome': ent.get('nome', '')})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@app.route('/entregas/<entrega_id>/remover', methods=['POST'])
def remover_pedido_entrega(entrega_id):
    """Remove UM pedido de uma entrega. Body: {pedidoId}. O pedido volta a
    pendente. Se a entrega ficar vazia, é apagada."""
    body = request.get_json(silent=True) or {}
    pid = (body.get('pedidoId') or '').strip()
    if not pid:
        return jsonify({'erro': 'pedidoId ausente'}), 400
    try:
        ent = carregar_entrega(entrega_id)
        if not ent:
            return jsonify({'erro': 'Entrega não encontrada'}), 404
        ids = [x for x in ent.get('pedidoIds', []) if x != pid]
        # devolve o pedido para pendente
        idx = {r['id']: r for r in listar_romaneios()}
        r = idx.get(pid)
        if r and r.get('status') == 'em_rota':
            r['status'] = 'pendente'
            r.pop('entregaId', None)
            r.pop('entregaNome', None)
            salvar_romaneio(pid, r)
        # entrega vazia -> apaga; senão salva
        if not ids:
            deletar_entrega(entrega_id)
            return jsonify({'ok': True, 'entregaVazia': True})
        ent['pedidoIds'] = ids
        salvar_entrega(entrega_id, ent)
        return jsonify({'ok': True, 'entregaVazia': False})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@app.route('/entregas/<entrega_id>/fase', methods=['POST'])
def set_fase_entrega(entrega_id):
    """Alterna a fase de uma entrega: 'planejamento' <-> 'em_rota'.
    'em_rota' = despachada (saiu da fábrica) -> deixa de contar no /embarque.
    Os pedidos continuam com status 'em_rota' nas duas fases (mapa/pedidos
    não mudam); a fase vive só na entrega."""
    body = request.get_json(silent=True) or {}
    nova = (body.get('fase') or '').strip()
    if nova not in ('planejamento', 'em_rota'):
        return jsonify({'erro': 'fase invalida'}), 400
    try:
        ent = carregar_entrega(entrega_id)
        if not ent:
            return jsonify({'erro': 'Entrega não encontrada'}), 404
        ent['fase'] = nova
        carimbo = datetime.datetime.utcnow().isoformat()
        if nova == 'em_rota':
            ent['despachadaEm'] = carimbo
        else:
            ent.pop('despachadaEm', None)
        salvar_entrega(entrega_id, ent)
        return jsonify({'ok': True, 'fase': nova})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


def _empresas_da_filial(fd, dados):
    """Empresas de faturamento presentes nos itens da filial (1=Indústria,
    2=Distribuidora). Ex.: [2] só distribuidora; [1, 2] pedido dividido."""
    return sorted({(i.get('empresa') or dados.get('empresa', 2)) for i in fd.get('itens', [])})


def _persistir_arquivos_romaneio(rid, dados, fd, emps, logo_bytes=None):
    """Salva o PDF e o Excel do romaneio. Se a filial tem as DUAS empresas,
    salva um par por empresa (ids sufixados '{rid}__e{N}', com empresa_override)
    — pra bater com o split do index e servir NF separada por emitente. Com uma
    empresa só, salva um arquivo único ('{rid}'), como antes. O front decide se
    baixa 1 ou 2 arquivos pelo campo 'empresas' do romaneio."""
    try:
        if len(emps) > 1:
            for e in emps:
                salvar_pedido_pdf(f'{rid}__e{e}', gerar_pdf({**dados, 'filiais': [fd]}, empresa_override=e, logo_bytes=logo_bytes))
        else:
            salvar_pedido_pdf(rid, gerar_pdf({**dados, 'filiais': [fd]}, logo_bytes=logo_bytes))
    except Exception as _e:
        print(f'[WARN] falha ao salvar PDF do romaneio {rid}: {_e}')
    try:
        if len(emps) > 1:
            for e in emps:
                salvar_pedido_excel(f'{rid}__e{e}', gerar_excel({**dados, 'filiais': [fd]}, empresa_override=e))
        else:
            salvar_pedido_excel(rid, gerar_excel({**dados, 'filiais': [fd]}))
    except Exception as _e:
        print(f'[WARN] falha ao salvar Excel do romaneio {rid}: {_e}')


@app.route('/romaneio-pdf/<rid>')
def romaneio_pdf(rid):
    """Serve inline o PDF da filial associado a um romaneio (abre no navegador)."""
    try:
        b = carregar_pedido_pdf(rid)
        if b is None:
            return jsonify({'erro': 'PDF não encontrado'}), 404
        return send_file(io.BytesIO(b), mimetype='application/pdf')
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/romaneio-excel/<rid>')
def romaneio_excel(rid):
    """Baixa o Excel da filial associado a um romaneio."""
    try:
        b = carregar_pedido_excel(rid)
        if b is None:
            return jsonify({'erro': 'Excel não encontrado (pedido antigo, sem Excel salvo)'}), 404
        return send_file(io.BytesIO(b),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f'pedido_{rid}.xlsx')
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/geofences')
def get_geofences():
    """Lista as zonas (geofences) salvas para o mapa."""
    try:
        return jsonify(listar_geofences())
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/geofence', methods=['POST'])
def post_geofence():
    """Cria ou atualiza uma zona. Se vier 'id', sobrescreve (edição de geometria/nome/cor)."""
    try:
        body = request.get_json(force=True) or {}
        geojson = body.get('geojson')
        if not geojson:
            return jsonify({'erro': 'geojson ausente'}), 400
        gid = body.get('id') or f"zona_{datetime.datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
        dados = {
            'id': gid,
            'nome': body.get('nome', gid),
            'cor': body.get('cor', '#8B1C1C'),
            'geojson': geojson,
        }
        salvar_geofence(gid, dados)
        return jsonify({'ok': True, 'geofence': dados})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/geofence/<geofence_id>', methods=['DELETE'])
def del_geofence(geofence_id):
    """Remove uma zona pelo id."""
    try:
        ok = deletar_geofence(geofence_id)
        if ok:
            return jsonify({'ok': True})
        return jsonify({'erro': 'Geofence não encontrada'}), 404
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# ─────────────────────────────────────────────────────────────────────
# AUTENTICAÇÃO
# ─────────────────────────────────────────────────────────────────────
def _sessao_atual():
    """Lê o token do header Authorization: Bearer <token> e valida."""
    auth = request.headers.get('Authorization', '')
    token = auth[7:].strip() if auth.startswith('Bearer ') else ''
    return (token, validar_sessao(token)) if token else ('', None)


def _exige_admin():
    _, s = _sessao_atual()
    return s if (s and s.get('papel') == 'admin') else None


@app.route('/login', methods=['POST'])
def login():
    body = request.get_json(silent=True) or {}
    usuario = (body.get('usuario') or '').strip()
    senha = body.get('senha') or ''
    if not usuario or not senha:
        return jsonify({'erro': 'Informe usuário e senha'}), 400
    users = carregar_usuarios()
    u = next((x for x in users if str(x.get('usuario', '')).lower() == usuario.lower()), None)
    ok = bool(u and u.get('ativo', True) and verifica_senha(senha, u.get('senhaHash', '')))
    try: registrar_login(usuario, ok)
    except Exception: pass
    if not ok:
        return jsonify({'erro': 'Usuário ou senha inválidos, ou usuário inativo'}), 401
    token = criar_sessao(u['usuario'], u.get('papel', 'operador'), u.get('codRepresentante', ''))
    return jsonify({'token': token, 'usuario': u['usuario'],
                    'papel': u.get('papel', 'operador'),
                    'nome': u.get('nome', u['usuario'])})


@app.route('/logout', methods=['POST'])
def logout():
    token, _ = _sessao_atual()
    encerrar_sessao(token)
    return jsonify({'ok': True})


@app.route('/me')
def me():
    _, s = _sessao_atual()
    if not s:
        return jsonify({'erro': 'sessão inválida'}), 401
    return jsonify({'usuario': s['usuario'], 'papel': s['papel'],
                    'codRepresentante': s.get('codRepresentante', '')})


@app.route('/admin/usuarios', methods=['GET'])
def admin_listar_usuarios():
    if not _exige_admin():
        return jsonify({'erro': 'acesso restrito'}), 403
    users = carregar_usuarios()
    # nunca devolve o hash da senha
    return jsonify([{k: v for k, v in u.items() if k != 'senhaHash'} for u in users])


@app.route('/admin/usuarios', methods=['POST'])
def admin_salvar_usuario():
    """Cria ou atualiza um usuário. Body: {usuario, nome, papel, ativo,
    codRepresentante, senha?}. Se 'senha' vier, gera novo hash; senão mantém."""
    if not _exige_admin():
        return jsonify({'erro': 'acesso restrito'}), 403
    body = request.get_json(silent=True) or {}
    usuario = (body.get('usuario') or '').strip()
    if not usuario:
        return jsonify({'erro': 'usuário obrigatório'}), 400
    users = carregar_usuarios()
    u = next((x for x in users if str(x.get('usuario', '')).lower() == usuario.lower()), None)
    novo = u is None
    if novo:
        u = {'usuario': usuario, 'senhaHash': ''}
        users.append(u)
    u['nome'] = body.get('nome', u.get('nome', usuario))
    u['papel'] = body.get('papel', u.get('papel', 'operador'))
    u['ativo'] = bool(body.get('ativo', u.get('ativo', True)))
    u['codRepresentante'] = body.get('codRepresentante', u.get('codRepresentante', ''))
    if body.get('senha'):
        u['senhaHash'] = hash_senha(body['senha'])
    if novo and not u['senhaHash']:
        return jsonify({'erro': 'defina uma senha para o novo usuário'}), 400
    salvar_usuarios(users)
    return jsonify({'ok': True, 'novo': novo})


@app.route('/health')
def health():
    # DIAGNÓSTICO TEMPORÁRIO (18/06): captura erro por cliente em vez de
    # deixar uma exceção silenciosa esconder o motivo de um perfil não
    # aparecer. Reverter pra versão simples depois de identificar a causa.
    perfis = {}
    erros = {}
    for c in {**CLIENTES, **CLIENTES_MANUAIS}:
        try:
            if perfil_existe(c):
                perfis[c] = perfil_filename(c)
        except Exception as e:
            erros[c] = f'{type(e).__name__}: {e}'
    # Lista de clientes disponíveis (PDF + manuais) para o frontend
    clientes_info = {}
    for cid, cdata in CLIENTES.items():
        clientes_info[cid] = {
            'nome': cdata['nome'],
            'tipo': 'proprio',
            'manual': False,
        }
    for cid, cdata in CLIENTES_MANUAIS.items():
        clientes_info[cid] = {
            'nome': cdata['nome'],
            'tipo': 'manual',
            'manual': True,
            'multiFilial': cdata.get('multiFilial', True),
        }

    resp = {'status': 'ok', 'perfis': perfis, 'clientes': clientes_info}
    if erros:
        resp['erros'] = erros
    return jsonify(resp)


def _validar_perfil(cliente, perfil_bytes):
    """Confere o perfil no momento do upload e devolve uma lista de AVISOS
    (não bloqueia o salvamento; é para corrigir o perfil na fonte). Foca no
    que faz o pedido sumir/ficar torto: filial sem lat/lng (não plota no mapa),
    tabela de filiais vazia e perfil sem produtos."""
    avisos = []
    try:
        _, produtos = ler_perfil(perfil_bytes)
        filiais = ler_filiais(perfil_bytes)
        multi = _cliente_meta(cliente).get('multiFilial', True) if _cliente_meta(cliente) else True
        if not produtos:
            avisos.append('O perfil não tem nenhum produto cadastrado (coluna C a partir da linha 8).')
        if multi and not filiais:
            avisos.append('A tabela de filiais (colunas M–T) está vazia — os pedidos não vão casar filial/região/coordenada.')
        for cnpj, info in filiais.items():
            nome = info.get('nome') or f'CNPJ {cnpj}'
            if info.get('lat') is None or info.get('lng') is None:
                avisos.append(f'Filial "{nome}" sem lat/lng (colunas S e T) — o pedido entra na lista, mas NÃO aparece no mapa.')
    except Exception as e:
        avisos.append(f'Não consegui validar o perfil por completo: {e}')
    return avisos


@app.route('/perfil/<cliente>', methods=['POST'])
def upload_perfil(cliente):
    """Salva ou atualiza o perfil de um cliente no servidor. Valida e devolve
    'avisos' (ex.: filial sem coordenada) para o front mostrar na tela."""
    if cliente not in CLIENTES and cliente not in CLIENTES_MANUAIS:
        return jsonify({'erro': f'Cliente inválido: {cliente}'}), 400
    f = request.files.get('perfil')
    if not f:
        return jsonify({'erro': 'Envie o arquivo perfil'}), 400
    filename = f.filename or ''
    perfil_bytes = f.read()
    salvar_perfil(cliente, perfil_bytes, filename)
    avisos = _validar_perfil(cliente, perfil_bytes)
    return jsonify({'ok': True, 'cliente': cliente, 'filename': filename,
                    'mensagem': 'Perfil salvo com sucesso', 'avisos': avisos})


@app.route('/master', methods=['POST'])
def upload_master():
    """Salva/atualiza a tabela mestra de produtos (MASTER.xlsx) e recarrega o cache."""
    f = request.files.get('master') or request.files.get('file')
    if not f:
        return jsonify({'erro': 'Envie o arquivo master'}), 400
    try:
        salvar_master(f.read())
        n = len(master.recarregar())
        return jsonify({'ok': True, 'codigos': n, 'mensagem': f'Tabela mestra salva ({n} códigos)'})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/master/status')
def master_status():
    """Informa se a tabela mestra existe e quantos códigos mapeia."""
    try:
        existe = master_existe()
        return jsonify({'existe': existe, 'codigos': len(master.get_mapa()) if existe else 0})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


@app.route('/logo/<cliente>')
def logo(cliente):
    """Retorna a logo extraída do perfil Excel do cliente."""
    if not perfil_existe(cliente):
        return jsonify({'erro': 'Perfil não encontrado'}), 404
    try:
        perfil_bytes = carregar_perfil_bytes(cliente)
        wb = openpyxl.load_workbook(io.BytesIO(perfil_bytes))
        ws = wb[wb.sheetnames[0]]
        if not ws._images:
            return jsonify({'erro': 'Sem imagem no perfil'}), 404
        img = ws._images[0]
        img.ref.seek(0)
        data = img.ref.read()
        return send_file(io.BytesIO(data), mimetype='image/png')
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


def _fmt_data_prog(iso):
    """Converte '2026-07-10T08:00:00' -> '10/07/2026' para o campo Data Entrega."""
    if not iso:
        return ''
    try:
        return datetime.datetime.fromisoformat(iso).strftime('%d/%m/%Y')
    except Exception:
        try:
            return datetime.fromisoformat(iso).strftime('%d/%m/%Y')
        except Exception:
            return ''


@app.route('/processar', methods=['POST'])
def processar():
    try:
        perfil_file = request.files.get('perfil')
        pedido_file = request.files.get('pedido')
        cliente = request.form.get('cliente', 'dom_atacarejo')
        data_prog = (request.form.get('dataEntregaProgramada') or '').strip() or None

        if not pedido_file:
            return jsonify({'erro': 'Envie o pedido'}), 400

        if cliente not in CLIENTES:
            return jsonify({'erro': f'Cliente {cliente} não implementado'}), 400

        # Perfil: usa o enviado agora (e salva) ou o salvo no servidor
        if perfil_file:
            perfil_bytes = perfil_file.read()
            salvar_perfil(cliente, perfil_bytes, perfil_file.filename)
        elif perfil_existe(cliente):
            perfil_bytes = carregar_perfil_bytes(cliente)
        else:
            return jsonify({'erro': f'Nenhum perfil disponível para {cliente}. Faça upload do perfil primeiro.'}), 400

        meta, produtos = ler_perfil(perfil_bytes)
        filiais_map = ler_filiais(perfil_bytes)
        pdf_bytes = pedido_file.read()

        parse_fn = CLIENTES[cliente]['parse']
        filiais = parse_fn(pdf_bytes, produtos)
        if cliente == 'assai' and filiais:
            meta['empresa'] = filiais[0]['empresa']

        if not filiais:
            return jsonify({'erro': 'Nenhuma filial encontrada no pedido'}), 400

        # Enriquecer cada filial com nome oficial e número, buscando pelo CNPJ
        # (regra única para todos os clientes: CNPJ é o dado mais confiável)
        for fd in filiais:
            nome_oficial, num_filial = buscar_filial(fd.get('cnpj', ''), filiais_map)
            if nome_oficial:
                fd['filial'] = nome_oficial
            if num_filial is not None:
                fd['numFilial'] = num_filial
            # Enriquecer com lat/lng do Perfil
            cnpj_norm = fd.get('cnpj', '').replace('.','').replace('/','').replace('-','')
            if cnpj_norm in filiais_map:
                fd['lat'] = filiais_map[cnpj_norm].get('lat')
                fd['lng'] = filiais_map[cnpj_norm].get('lng')
                fd['regiao'] = filiais_map[cnpj_norm].get('regiao')

        dados = {**meta, 'filiais': filiais, 'clienteNome': CLIENTES[cliente]['nome']}

        # Extrair logo do perfil para o PDF
        logo_bytes = None
        try:
            wb_logo = openpyxl.load_workbook(io.BytesIO(perfil_bytes))
            ws_logo = wb_logo[wb_logo.sheetnames[0]]
            if ws_logo._images:
                ws_logo._images[0].ref.seek(0)
                logo_bytes = ws_logo._images[0].ref.read()
        except Exception:
            pass

        arquivos, eb_simples, pb_simples, split = _gerar_arquivos_por_empresa(dados, filiais, logo_bytes=logo_bytes)

        # Salvar o romaneio de cada filial. O romaneio alimenta a LISTA e o
        # mapa; salvamos SEMPRE que houver itens, mesmo sem lat/lng (aí o
        # pedido aparece na lista e o mapa só não plota o pino — o mapa já
        # ignora romaneio sem coordenada). Antes, sem coordenada o pedido
        # sumia da lista também (bug: filial sem lat/lng no perfil).
        from datetime import datetime
        for fd in filiais:
            lat = fd.get('lat')
            lng = fd.get('lng')
            its = fd.get('itens', [])
            if not its:
                continue
            emps = _empresas_da_filial(fd, dados)
            ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            filial_slug = _re.sub(r'[^a-z0-9]', '_', fd['filial'].lower())
            rid = f"{cliente}_{filial_slug}_{ts}"
            salvar_romaneio(rid, {
                'id': rid,
                'cliente': cliente,
                'clienteNome': CLIENTES[cliente]['nome'],
                'filial': fd['filial'],
                'numero': fd.get('numFilial', ''),
                'regiao': fd.get('regiao', ''),
                'cnpj': fd.get('cnpj', ''),
                'lat': lat,
                'lng': lng,
                'dataPedido': fd.get('dataPedido', fd.get('dataEmissao', '')),
                'dataGeracao': datetime.utcnow().isoformat(),
                'kgPlanejados': round(sum(float(i.get('kgPlanejados', 0)) for i in its), 1),
                'itens': [{'cod': str(i.get('codInterno') or '').strip(), 'nome': str(i.get('nomeProduto') or ''), 'kg': round(_kg_pdf(i), 3)} for i in its],
                'pedidoNum': fd.get('pedidoNum', ''),
                'empresas': emps,
                'dataEntregaProgramada': data_prog,
            })
            # Se o pedido foi programado, usa a data programada no campo Data Entrega
            if data_prog:
                fd['dataEntrega'] = _fmt_data_prog(data_prog)
            # Persistir PDF+Excel do romaneio (separados por empresa se for pedido dividido)
            _persistir_arquivos_romaneio(rid, dados, fd, emps, logo_bytes)

        todos_itens = [i for f in filiais for i in f['itens']]
        return jsonify({
            'ok': True,
            'split': split,
            'filiais': len(filiais),
            'itens': len(todos_itens),
            'totalKg': round(sum(i['kgPlanejados'] for i in todos_itens), 1),
            'totalValor': round(sum(i['valorPedido'] for i in todos_itens), 2),
            'resumo': [{'filial': f['filial'], 'pedidoNum': f.get('pedidoNum', ''),
                        'itens': len(f['itens']),
                        'kg': round(sum(i['kgPlanejados'] for i in f['itens']), 1),
                        'valor': round(sum(i['valorPedido'] for i in f['itens']), 2)}
                       for f in filiais],
            'arquivos': arquivos,
            # compatibilidade retroativa (caso simples) — reaproveita o que já foi gerado acima
            'excel': base64.b64encode(eb_simples).decode() if eb_simples is not None else '',
            'pdf': base64.b64encode(pb_simples).decode() if pb_simples is not None else '',
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


@app.route('/operadores/<cliente>')
def operadores_cliente(cliente):
    """Lista os operadores cadastrados no Perfil de um cliente do fluxo
    manual, pra alimentar o dropdown de quem está lançando o pedido."""
    if cliente not in CLIENTES_MANUAIS and cliente not in CLIENTES:
        return jsonify({'erro': f'Cliente {cliente} não encontrado'}), 400
    if not perfil_existe(cliente):
        return jsonify({'erro': 'Perfil não encontrado'}), 404
    perfil_bytes = carregar_perfil_bytes(cliente)
    return jsonify({'operadores': ler_operadores(perfil_bytes)})


@app.route('/filiais/<cliente>')
def filiais_cliente(cliente):
    """Lista as filiais disponíveis pro popup de pedido manual.
    Clientes multiFilial=True: lê a tabela M:N:O:P:Q do Perfil (várias lojas).
    Clientes multiFilial=False (ex: Guanabara Central): não existe tabela —
    devolve uma única filial sintetizada a partir do cabeçalho do Perfil
    (CNPJ/Filial/Endereço únicos), pra manter o mesmo formato de resposta
    e o frontend não precisar de um caminho de código totalmente separado."""
    if cliente not in CLIENTES_MANUAIS and cliente not in CLIENTES:
        return jsonify({'erro': f'Cliente {cliente} não encontrado'}), 400
    if not perfil_existe(cliente):
        return jsonify({'erro': 'Perfil não encontrado'}), 404
    perfil_bytes = carregar_perfil_bytes(cliente)
    if _cliente_meta(cliente).get('multiFilial', True):
        filiais_map = ler_filiais(perfil_bytes)
        lista = [{'cnpj': cnpj, 'nome': info['nome'], 'numero': info['numero'],
                  'endereco': info['endereco'], 'cidade': info['cidade']}
                 for cnpj, info in filiais_map.items()]
    else:
        meta, _ = ler_perfil(perfil_bytes)
        lista = [{
            'cnpj': _normaliza_cnpj(meta.get('cnpjPerfil', '')),
            'nome': meta.get('filialPerfil') or meta.get('clienteNomePerfil') or _cliente_meta(cliente)['nome'],
            'numero': None,
            'endereco': meta.get('enderecoPerfil', ''),
            'cidade': '',
        }]
    lista.sort(key=lambda f: f['nome'])
    return jsonify({'filiais': lista, 'multiFilial': _cliente_meta(cliente).get('multiFilial', True)})


@app.route('/produtos/<cliente>')
def produtos_cliente(cliente):
    """Lista os produtos cadastrados no Perfil de um cliente do fluxo
    manual, pra pré-popular as linhas do popup (nome, formato, embalagem
    e a unidade em que a quantidade deve ser digitada — cx ou kg)."""
    if cliente not in CLIENTES_MANUAIS and cliente not in CLIENTES:
        return jsonify({'erro': f'Cliente {cliente} não encontrado'}), 400
    if not perfil_existe(cliente):
        return jsonify({'erro': 'Perfil não encontrado'}), 404
    perfil_bytes = carregar_perfil_bytes(cliente)
    _, produtos = ler_perfil(perfil_bytes)
    lista = [{
        'index': i,
        'codInterno': p['codInterno'],
        'nomePerfil': p['nomePerfil'],
        'formato': p['formato'],
        'embalagem': p['embalagem'],
        'unidFat': p['unidFat'],
        'kgCx': p['kgCx'],
        'precoUnit': p['precoUnit'],
    } for i, p in enumerate(produtos)]
    return jsonify({'produtos': lista})


@app.route('/processar-manual', methods=['POST'])
def processar_manual():
    """Gera Excel+PDF a partir de um pedido lançado manualmente no popup
    (clientes sem PDF de pedido — pedidos por telefone/WhatsApp em formato
    livre). O operador escolhe a filial pelo nome e digita a quantidade de
    cada produto já pré-carregado do Perfil; nenhum parsing de texto livre
    é necessário, já que o produto e a filial vêm de seleção direta."""
    try:
        body = request.get_json(force=True) or {}
        data_prog = (body.get('dataEntregaProgramada') or '') or None
        cliente = body.get('cliente', '')
        operador = body.get('operador', '')
        filial_nome = body.get('filialNome', '')
        itens_form = body.get('itens', [])

        if cliente not in CLIENTES_MANUAIS and cliente not in CLIENTES:
            return jsonify({'erro': f'Cliente {cliente} não encontrado'}), 400
        if not perfil_existe(cliente):
            return jsonify({'erro': f'Nenhum perfil disponível para {cliente}'}), 400
        if not operador:
            return jsonify({'erro': 'Selecione o operador'}), 400

        multi_filial = _cliente_meta(cliente).get('multiFilial', True)
        if multi_filial and not filial_nome:
            return jsonify({'erro': 'Selecione a filial'}), 400

        perfil_bytes = carregar_perfil_bytes(cliente)
        meta, produtos = ler_perfil(perfil_bytes)
        operadores_validos = ler_operadores(perfil_bytes)

        if operador not in operadores_validos:
            return jsonify({'erro': f'Operador "{operador}" não cadastrado no perfil'}), 400

        if multi_filial:
            # encontra a filial selecionada pelo nome (a seleção é direta, não por CNPJ extraído de PDF)
            filiais_map = ler_filiais(perfil_bytes)
            cnpj_sel, filial_info = None, None
            for cnpj, info in filiais_map.items():
                if info['nome'] == filial_nome:
                    cnpj_sel, filial_info = cnpj, info
                    break
            if not filial_info:
                return jsonify({'erro': f'Filial "{filial_nome}" não encontrada no perfil'}), 400
        else:
            # cliente 'Central': CNPJ/Filial/Endereço únicos, vêm do cabeçalho do Perfil
            cnpj_sel = _normaliza_cnpj(meta.get('cnpjPerfil', ''))
            _central = ler_filiais(perfil_bytes).get(cnpj_sel, {})
            filial_info = {
                'nome': meta.get('filialPerfil') or meta.get('clienteNomePerfil') or _cliente_meta(cliente)['nome'],
                'numero': _central.get('numero'),
                'endereco': meta.get('enderecoPerfil', ''),
                'lat': _central.get('lat'),
                'lng': _central.get('lng'),
                'regiao': _central.get('regiao'),
            }

        itens = []
        for it in itens_form:
            qtde = it.get('quantidade')
            if not qtde:
                continue  # linha em branco, produto não pedido nessa ligação
            qtde = float(qtde)
            if qtde <= 0:
                continue
            idx = it.get('index')
            produto = produtos[idx] if isinstance(idx, int) and 0 <= idx < len(produtos) else None
            if not produto:
                continue
            kgCx = produto['kgCx']
            unidFat = produto['unidFat']
            kgPlan = qtde * kgCx if unidFat == 'cx' else qtde
            preco = produto['precoUnit']
            itens.append({
                'empresa': produto.get('empresa'),
                'codInterno': produto['codInterno'],
                'nomeProduto': produto['nomePerfil'],
                'formato': produto.get('formato', ''),
                'embalagem': produto.get('embalagem', ''),
                'kgCx': kgCx,
                'kgPlanejados': kgPlan,
                'nrCaixas': round(kgPlan / kgCx, 1) if kgCx else 0,
                'obs': produto.get('obs', ''),
                'qtdeMultipl': qtde if unidFat == 'cx' else kgPlan,
                'precoUnit': preco,
                'valorPedido': round(kgPlan * preco, 2),
                'precoSistema': preco,
                'unidFat': unidFat,
            })

        if not itens:
            return jsonify({'erro': 'Nenhum item com quantidade preenchida'}), 400

        agora = datetime.datetime.now()
        pedido_num = body.get('pedidoNum', '').strip() or f"MANUAL-{agora.strftime('%Y%m%d-%H%M%S')}"

        lat = filial_info.get('lat')
        lng = filial_info.get('lng')

        filiais = [{
            'filial': filial_info['nome'],
            'numFilial': filial_info['numero'],
            'pedidoNum': pedido_num,
            'cnpj': cnpj_sel,
            'endereco': filial_info['endereco'],
            'dataPedido': agora.strftime('%d/%m/%Y'),
            'dataEntrega': '',
            'condPgto': '',
            'solicitante': operador,
            'empresa': meta.get('empresa', 2),
            'regiao': filial_info.get('regiao'),
            'itens': itens,
            'lat': lat,
            'lng': lng,
        }]

        dados = {**meta, 'filiais': filiais, 'clienteNome': _cliente_meta(cliente)['nome']}

        # Extrair logo do perfil para o PDF (antes de salvar o romaneio, pois o
        # PDF do romaneio também precisa da logo)
        logo_bytes = None
        try:
            import openpyxl, io as _io
            wb_logo = openpyxl.load_workbook(_io.BytesIO(perfil_bytes))
            ws_logo = wb_logo[wb_logo.sheetnames[0]]
            if ws_logo._images:
                ws_logo._images[0].ref.seek(0)
                logo_bytes = ws_logo._images[0].ref.read()
        except Exception:
            pass

        # Salvar pin de mapa se tiver coordenadas
        if lat is not None and lng is not None:
            ts = datetime.datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            filial_slug = ''.join(c if c.isalnum() else '_' for c in filial_info['nome'].lower())
            rid = f"{cliente}_{filial_slug}_{ts}"
            salvar_romaneio(rid, {
                'id': rid,
                'cliente': cliente,
                'clienteNome': _cliente_meta(cliente)['nome'],
                'filial': filial_info['nome'],
                'numero': filial_info.get('numero'),
                'regiao': filial_info.get('regiao', ''),
                'cnpj': cnpj_sel,
                'lat': lat,
                'lng': lng,
                'dataPedido': agora.strftime('%d/%m/%Y'),
                'dataGeracao': datetime.datetime.utcnow().isoformat(),
                'dataEntregaProgramada': data_prog,
                'kgPlanejados': round(sum(float(i.get('kgPlanejados', 0)) for i in itens), 1),
                'itens': [{'cod': str(i.get('codInterno') or '').strip(), 'nome': str(i.get('nomeProduto') or ''), 'kg': round(_kg_pdf(i), 3)} for i in itens],
                'pedidoNum': pedido_num,
                'empresas': _empresas_da_filial(filiais[0], dados),
            })
            # Persistir PDF+Excel do romaneio (separados por empresa se dividido)
            if data_prog:
                filiais[0]['dataEntrega'] = _fmt_data_prog(data_prog)
            _persistir_arquivos_romaneio(rid, dados, filiais[0],
                                         _empresas_da_filial(filiais[0], dados), logo_bytes)
        else:
            print(f'[INFO] romaneio não salvo — lat={lat} lng={lng}')

        arquivos, eb_simples, pb_simples, split = _gerar_arquivos_por_empresa(dados, filiais, logo_bytes=logo_bytes)

        return jsonify({
            'ok': True,
            'split': split,
            'filiais': 1,
            'itens': len(itens),
            'totalKg': round(sum(i['kgPlanejados'] for i in itens), 1),
            'totalValor': round(sum(i['valorPedido'] for i in itens), 2),
            'pedidoNum': pedido_num,
            'resumo': [{'filial': filial_info['nome'], 'pedidoNum': pedido_num,
                        'itens': len(itens),
                        'kg': round(sum(i['kgPlanejados'] for i in itens), 1),
                        'valor': round(sum(i['valorPedido'] for i in itens), 2)}],
            'arquivos': arquivos,
            'excel': base64.b64encode(eb_simples).decode() if eb_simples is not None else '',
            'pdf': base64.b64encode(pb_simples).decode() if pb_simples is not None else '',
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500


if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)




