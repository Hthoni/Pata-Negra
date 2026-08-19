"""
Cliente Avulso — pedidos pontuais sem parser/perfil dedicado.

Em vez de um Perfil por cliente (produtos + filiais cadastrados), usa uma
única planilha de FATURAMENTO com os dados fixos por CNPJ (código do
cliente, razão social, código da condição de pagamento, vendedor + código,
endereço de entrega, região e coordenadas). O operador digita o CNPJ no
popup, o sistema completa esses dados automaticamente (incluindo o pino no
mapa), e monta o pedido linha a linha escolhendo o produto da tabela
MASTER (não existe perfil próprio pra esse "cliente", então não há preço
de referência nem alerta de divergência de preço pra esses pedidos).

Planilha de faturamento (aba única, uma linha por CNPJ) — layout real
(exportado do sistema de faturamento existente do cliente):
  Col A: Código Cliente (código interno do sistema de faturamento)
  Col B: CNPJ
  Col C: Razão Social
  Col D: Código da Condição de Pagamento
  Col E: Vendedor (nome)
  Col F: Código do Vendedor
  Col G: Endereço de Entrega
  Col H: Região
  Col I: Lat
  Col J: Long
Cabeçalho na linha 1; dados a partir da linha 2.

FIX (19/08/2026): adicionadas as colunas H (Região), I (Lat) e J (Long) —
sem elas, o romaneio do pedido avulso nunca ganhava pino no mapa nem
região de roteirização (main.py salvava lat=None/lng=None de propósito,
já que não havia de onde tirar essas coordenadas). Lat/Long tolera vírgula
decimal (padrão BR, ex.: "-22,8263973"), igual ao perfil.py normal.

FIX (18/08/2026, mantido): layout real não tem coluna de texto "Condição
de Pagamento" nem CNPJ na coluna A — a coluna A é "Código Cliente", CNPJ
é a B. Ver histórico do módulo se precisar do contexto completo.
"""
import io
import re
import openpyxl


def _normaliza_cnpj(cnpj):
    return re.sub(r'\D', '', str(cnpj or ''))


def _coord(v):
    """Lat/Lng tolerante a vírgula decimal (BR) ou ponto; vazio/ruim -> None
    (a filial fica sem pino no mapa em vez de derrubar a leitura inteira)."""
    if v is None:
        return None
    try:
        return float(str(v).strip().replace(',', '.'))
    except (ValueError, TypeError):
        return None


def ler_faturamento_avulso(xlsx_bytes):
    """Lê a planilha de faturamento avulso e devolve
    {cnpj_normalizado: {codigoCliente, cnpj, razaoSocial, codCondPgto,
                         vendedor, codVendedor, endereco, regiao, lat, lng}}."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or not row[1]:
            continue
        cnpj_raw = str(row[1]).strip()
        cnpj_norm = _normaliza_cnpj(cnpj_raw)
        if not cnpj_norm:
            continue
        out[cnpj_norm] = {
            'codigoCliente': str(row[0] or '').strip(),
            'cnpj': cnpj_raw,
            'razaoSocial': str(row[2] or '').strip() if len(row) > 2 else '',
            'codCondPgto': str(row[3] or '').strip() if len(row) > 3 and row[3] is not None else '',
            'vendedor': str(row[4] or '').strip() if len(row) > 4 and row[4] is not None else '',
            'codVendedor': str(row[5] or '').strip() if len(row) > 5 and row[5] is not None else '',
            'endereco': str(row[6] or '').strip() if len(row) > 6 else '',
            'regiao': str(row[7] or '').strip() if len(row) > 7 and row[7] is not None else '',
            'lat': _coord(row[8]) if len(row) > 8 else None,
            'lng': _coord(row[9]) if len(row) > 9 else None,
        }
    return out


def calc_item_avulso(qtde, unidade, peso_unit_kg, preco_unit):
    """kg e valor de um item avulso, a partir da unidade escolhida pelo
    operador (kg -> qtde já é kg; pct/cx -> qtde x peso da unidade em kg).
    Preço sempre na MESMA unidade escolhida (preço por kg se unidade=kg,
    preço por caixa se unidade=cx etc) — valor = qtde x preço, sempre."""
    qtde = float(qtde or 0)
    preco_unit = float(preco_unit or 0)
    if unidade == 'kg':
        kg = qtde
    else:
        kg = qtde * float(peso_unit_kg or 0)
    valor = round(qtde * preco_unit, 2)
    return round(kg, 3), valor
