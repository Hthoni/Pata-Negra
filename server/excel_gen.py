"""
Gerador do Excel de upload — uma aba por filial, com fórmulas vivas
e formatação completa (cores, bordas, mesclagens).
"""
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


def _thin():
    return Side(style='thin')


def _brd(t=0, b=0, l=0, r=0):
    return Border(
        top=_thin() if t else Side(), bottom=_thin() if b else Side(),
        left=_thin() if l else Side(), right=_thin() if r else Side())


def _fill(rgb):
    return PatternFill(fill_type='solid', fgColor=rgb)


def _fnt(bold=False, sz=9, color='FF000000'):
    return Font(bold=bold, size=sz, color=color)


def _aln(h='left', wrap=False):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)


BALL = _brd(1, 1, 1, 1)
BBOT = _brd(b=1)
BG_TIT = _fill('FFE0E0E0')
BG_SUB = _fill('FFF0F0F0')
BG_META = _fill('FFF8F8F8')
BG_HDR = _fill('FFD0D0D0')
BG_AZL = _fill('FF2980B9')
BG_AZLT = _fill('FFEBF5FB')
BG_CZ = _fill('FFE8E8E8')
BG_PAR = _fill('FFF5F5F5')
BG_TOTV = _fill('FFD5F5E3')
F_TIT = _fnt(True, 11)
F_SUB = _fnt(True, 10)
F_MB = _fnt(True, 9)
F_M = _fnt(sz=9)
F_HW = _fnt(True, 9, 'FFFFFFFF')
F_HB = _fnt(True, 9)
F_IT = _fnt(sz=9)
F_TOT = _fnt(True, 9)
F_NT = _fnt(sz=8, color='FF666666')
F_NR = _fnt(sz=8, color='FFC0392B')
F_CX = _fnt(True, 9, 'FF854F0B')
F_DIF = _fnt(True, 9, 'FFCC0000')

FMT_MONEY = '#.##0,00'
FMT_NUM  = '#.##0,0'
FMT_INT  = '0'       # número inteiro (ex: nº caixas)


def _ap(cell, val=None, fn=None, bg=None, bo=None, al=None, fmt=None):
    if val is not None:
        cell.value = val
    if fn:
        cell.font = fn
    if bg:
        cell.fill = bg
    if bo:
        cell.border = bo
    if al:
        cell.alignment = al
    if fmt:
        cell.number_format = fmt


def gerar_excel(dados, empresa_override=None):
    """Gera o Excel completo. Se empresa_override for passado, filtra
    apenas os itens daquela empresa (usado no modo split)."""
    emp = empresa_override if empresa_override else dados.get('empresa', 2)
    cv = str(dados.get('codVend', ''))
    cc = str(dados.get('codCond', ''))
    vend = dados.get('vendedor', '')
    tel = dados.get('telefone', '')
    cli = dados.get('clienteNome', '')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for fd in dados['filiais']:
        if empresa_override:
            its = [it for it in fd['itens']
                   if (it.get('empresa') or empresa_override) == empresa_override]
        else:
            its = fd['itens']
        if not its:
            continue
        n = len(its)
        emp_fd = empresa_override if empresa_override else fd.get('empresa', emp)
        tit_fd = 'PEDIDO PATA NEGRA DISTRIBUIDORA' if emp_fd == 2 else 'PEDIDO INDÚSTRIA PATA NEGRA'
        numPed = re.sub(r'[/\\*?\[\]:]', '-', fd.get('pedidoNum', '')).strip().strip('-')
        nomeAba = ((numPed + ' - ' + fd['filial']) if numPed else fd['filial'])[:31]
        ws = wb.create_sheet(nomeAba)

        for i, w in enumerate([4, 10, 38, 12, 7, 12, 12, 9, 14, 2, 11, 12, 14, 12, 10, 8], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.merge_cells('A1:I1')
        _ap(ws['A1'], tit_fd, F_TIT, BG_TIT, al=_aln('center'))
        ws.merge_cells('A2:I2')
        subtitulo = (cli + ' — ' + fd['filial']) if cli else fd['filial']
        _ap(ws['A2'], subtitulo, F_SUB, BG_SUB, al=_aln('center'))

        for r in [3, 4, 6, 7]:
            ws.merge_cells(f'A{r}:B{r}')
            ws.merge_cells(f'C{r}:F{r}')
            ws.merge_cells(f'H{r}:I{r}')
        # Linha 5 não mescla C:F — fica C:D (nome filial) + E (label) + F (número)
        ws.merge_cells('A5:B5')
        ws.merge_cells('C5:D5')
        ws.merge_cells('H5:I5')

        _ap(ws['A3'], 'Pedido Nº:', F_MB, BG_META, al=_aln('left'))
        _ap(ws['C3'], fd.get('pedidoNum', ''), F_M, BG_META)
        _ap(ws['G3'], 'Data Pedido:', F_MB, BG_META, al=_aln('left'))
        _ap(ws['H3'], fd.get('dataPedido', ''), F_M, BG_META, al=_aln('left'))

        _ap(ws['A4'], 'CNPJ:', F_MB, BG_META, al=_aln('left'))
        _ap(ws['C4'], fd.get('cnpj', ''), F_M, BG_META)
        _ap(ws['G4'], 'Data Entrega:', F_MB, BG_META, al=_aln('left'))
        _ap(ws['H4'], fd.get('dataEntrega', ''), F_M, BG_META, al=_aln('left'))

        _ap(ws['A5'], 'Filial:', F_MB, BG_META, al=_aln('left'))
        _ap(ws['C5'], fd['filial'], F_M, BG_META)
        _ap(ws['E5'], 'Núm. Filial:', F_MB, BG_META, al=_aln('left'))
        _ap(ws['F5'], fd.get('numFilial', ''), F_M, BG_META, al=_aln('center'))
        _ap(ws['G5'], 'Solicitante:', F_MB, BG_META, al=_aln('left'))
        _ap(ws['H5'], fd.get('solicitante', ''), F_M, BG_META, al=_aln('left'))
        _ap(ws['K5'], 'Código vend.', F_HW, BG_AZL, BALL, _aln('center', True))
        _ap(ws['L5'], 'código cond', F_HW, BG_AZL, BALL, _aln('center', True))
        _ap(ws['M5'], 'empresa', F_HW, BG_AZL, BALL, _aln('center', True))

        _ap(ws['A6'], 'Endereço:', F_MB, BG_META, al=_aln('left'))
        _ap(ws['C6'], fd.get('endereco', ''), F_M, BG_META)
        _ap(ws['G6'], 'Vendedor:', F_MB, BG_META, al=_aln('left'))
        _ap(ws['H6'], vend, F_M, BG_META, al=_aln('left'))
        ws['K6'].value = int(cv) if cv.isdigit() else cv
        ws['L6'].value = int(cc) if cc.isdigit() else cc
        ws['M6'].value = emp_fd

        _ap(ws['A7'], 'Cond. Pgto.:', F_MB, BG_META, BBOT, _aln('left'))
        _ap(ws['C7'], fd.get('condPgto', ''), F_M, BG_META, BBOT)
        for col in ['B7', 'D7', 'E7', 'F7']:
            ws[col].border = BBOT
        _ap(ws['G7'], 'Tel. Vendedor:', F_MB, BG_META, al=_aln('left'))
        _ap(ws['H7'], tel, F_M, BG_META, al=_aln('left'))

        for col, h, az in [('A', '#', 0), ('B', 'Cód.\nInterno', 0), ('C', 'Nome Produto\nno Cliente', 0),
                            ('D', 'Formato', 0), ('E', 'Caixa', 0), ('F', 'Kg\nPlanejados', 0),
                            ('G', 'Kg\nEmbarcados', 0), ('H', 'Nº\nCaixas', 0), ('I', 'Obs.', 0),
                            ('K', 'Qtde\nMultipl.', 1), ('L', 'Preço Unit.\n(R$)', 1),
                            ('M', 'Valor Pedido\n(R$)', 1), ('N', 'Preço\nSistema', 1),
                            ('O', 'Dif.\nPreço', 1), ('P', 'Unid.\nfat.', 1)]:
            _ap(ws[f'{col}8'], h, F_HW if az else F_HB, BG_AZL if az else BG_HDR, BALL, _aln('center', True))
        ws.row_dimensions[8].height = 28

        for idx, it in enumerate(its):
            r = idx + 9
            par = (idx % 2 == 1)
            bg = BG_PAR if par else None
            isCx = (it.get('unidFat', 'kg') == 'cx')
            kgcx = it.get('kgCx', 20)

            def c(col):
                return ws[f'{col}{r}']

            _ap(c('A'), idx + 1, F_IT, bg, BALL, _aln('center'))
            _ap(c('B'), it.get('codInterno'), F_IT, bg, BALL, _aln('center'))
            _ap(c('C'), it.get('nomeProduto', ''), F_IT, bg, BALL, _aln('left'))
            _ap(c('D'), it.get('formato') or None, F_IT, bg, BALL, _aln('center'))
            _ap(c('E'), it.get('embalagem', ''), F_IT, bg, BALL, _aln('center'))
            _ap(c('F'), it.get('kgPlanejados'), F_IT, bg, BALL, _aln('center'), FMT_NUM)
            _ap(c('G'), None, F_IT, BG_CZ, BALL, _aln('center'))
            if isCx:
                c('K').value = f'=IF(G{r}<>"",G{r}/{kgcx},F{r}/{kgcx})'
            else:
                c('K').value = f'=IF(G{r}<>"",G{r},F{r})'
            _ap(c('K'), None, F_IT, BG_AZLT, BALL, _aln('center'))
            c('H').value = f'=IF(G{r}<>"",G{r}/{kgcx},F{r}/{kgcx})'
            _ap(c('H'), None, F_IT, bg, BALL, _aln('center'), FMT_INT)
            _ap(c('I'), it.get('obs') or None, F_IT, bg, BALL, _aln('center'))
            _ap(c('L'), it.get('precoUnit'), F_IT, BG_AZLT, BALL, _aln('center'), FMT_MONEY)
            c('M').value = f'=L{r}*K{r}' if isCx else f'=IF(G{r}<>"",L{r}*G{r},L{r}*F{r})'
            _ap(c('M'), None, F_IT, BG_AZLT, BALL, _aln('center'), FMT_MONEY)
            _ap(c('N'), it.get('precoSistema') or 0, F_IT, BG_AZLT, BALL, _aln('center'), FMT_MONEY)
            c('O').value = f'=IF(AND(L{r}<>"",N{r}<>"",L{r}<>N{r}),L{r}-N{r},"")'
            _ap(c('O'), None, F_DIF, BG_AZLT, BALL, _aln('center'), FMT_MONEY)
            _ap(c('P'), it.get('unidFat', 'kg'), F_CX if isCx else F_IT, BG_AZLT, BALL, _aln('center'))

        ws.conditional_formatting.add(
            f'O9:O{n+8}',
            FormulaRule(formula=['O9=""'], font=Font(bold=False, size=9, color='FF000000'))
        )

        rT = n + 9
        r1 = 9
        r2 = n + 8
        ws.merge_cells(f'A{rT}:E{rT}')
        _ap(ws[f'A{rT}'], f'TOTAL  —  {n} itens', F_TOT, None, _brd(1, 1, 1, 0), _aln('left'))
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{rT}'].border = _brd(1, 1, 0, 0)
        for col in ['F', 'G', 'H']:
            ws[f'{col}{rT}'].value = f'=SUM({col}{r1}:{col}{r2})'
            _ap(ws[f'{col}{rT}'], None, F_TOT, BG_HDR, BALL, _aln('center'), FMT_NUM)
        ws[f'M{rT}'].value = f'=SUM(M{r1}:M{r2})'
        _ap(ws[f'M{rT}'], None, F_TOT, BG_TOTV, BALL, _aln('center'), FMT_MONEY)

        rI = rT + 2
        ws.merge_cells(f'A{rI}:I{rI}')
        ws.merge_cells(f'K{rI}:P{rI}')
        _ap(ws[f'A{rI}'], f'Imprimir: selecionar A1:I{rT-1} → Imprimir Seleção', F_NT, al=_aln('left'))
        _ap(ws[f'K{rI}'], 'Col K=qtde faturamento | Col P=unidade | Dif. Preço vermelho=revisar', F_NR, al=_aln('left'))

    wsr = wb.create_sheet('RESUMO GERAL')
    for i, w in enumerate([22, 12, 22, 6, 12, 14], 1):
        wsr.column_dimensions[get_column_letter(i)].width = w
    for ci, h in enumerate(['Filial', 'Pedido Nº', 'CNPJ', 'Itens', 'Kg Plan.', 'Valor (R$)'], 1):
        _ap(wsr.cell(1, ci), h, F_HB, BG_HDR, BALL, _aln('center'))
    skg = sv = sit = 0
    for ri, fd in enumerate(dados['filiais'], 2):
        its2 = fd['itens']
        kg = sum(i.get('kgPlanejados', 0) for i in its2)
        val = sum(i.get('valorPedido', 0) for i in its2)
        skg += kg
        sv += val
        sit += len(its2)
        par = (ri % 2 == 0)
        bg2 = BG_PAR if par else None
        for ci, v in enumerate([fd['filial'], fd.get('pedidoNum', ''), fd.get('cnpj', ''),
                                 len(its2), round(kg, 1), round(val, 2)], 1):
            fmt2 = FMT_MONEY if ci == 6 else (FMT_NUM if ci == 5 else None)
            _ap(wsr.cell(ri, ci), v, F_IT, bg2, BALL, fmt=fmt2)
    rTR = len(dados['filiais']) + 2
    for ci, v in enumerate(['TOTAL', '', '', sit, round(skg, 1), round(sv, 2)], 1):
        fmt2 = FMT_MONEY if ci == 6 else (FMT_NUM if ci == 5 else None)
        _ap(wsr.cell(rTR, ci), v, F_TOT, BG_HDR, BALL, fmt=fmt2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
        
