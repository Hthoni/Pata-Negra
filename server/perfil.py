"""
Leitura do Perfil Excel de cada cliente e algoritmo de matching
de produtos do pedido (PDF) com produtos cadastrados no perfil.
"""
import io
import re
import unicodedata
import openpyxl

CNPJ_DISTRIBUIDORA = '56.423.719'
CNPJ_INDUSTRIA = '10.171.633'


def _normaliza_nome(s):
    """Normaliza um nome de produto para comparação: sem acento, minúsculas,
    espaços múltiplos colapsados, sem espaço nas pontas. Usado só para achar
    o produto certo — NÃO decide a grafia final (essa continua vindo do
    perfil, coluna C)."""
    s = str(s or '')
    s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('ASCII')
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return s


def match_perfil(nome, produtos):
    """Encontra o produto do perfil cujo nome bate EXATAMENTE (tolerando só
    maiúscula/minúscula, acentuação e espaços extras) com o nome do PDF.

    FIX (01/08/2026): antes fazia matching APROXIMADO — substring ('n in a'
    / 'a in n') ou contagem de palavras em comum — o que às vezes casava o
    pedido com um produto PARECIDO mas ERRADO quando o nome do cliente vinha
    um pouco diferente do cadastrado (ex.: faltando "DEF", ou um nome mais
    curto que por acaso é substring de outro). Isso é silencioso e perigoso:
    o pedido processa normal, só que o item físico errado. Agora só casa por
    igualdade exata normalizada; sem correspondência exata, retorna None —
    e processar_item() interrompe o pedido com um erro claro em vez de
    seguir adiante com um "SEM MASTER" ou uma aproximação arriscada."""
    alvo = _normaliza_nome(nome)
    if not alvo:
        return None
    for p in produtos:
        a = _normaliza_nome(p.get('nomePerfil', ''))
        if a and a == alvo:
            return p
    return None


def ler_perfil(perfil_bytes):
    """Lê o Perfil Excel e retorna (meta, produtos).
    meta: dados do cabeçalho (vendedor, telefone, códigos).
    produtos: lista de dicts com os produtos cadastrados, incluindo
    a coluna A (Fat.) que define a empresa de faturamento (1=Indústria, 2=Distribuidora)."""
    wb_p = openpyxl.load_workbook(io.BytesIO(perfil_bytes), data_only=True)
    pws = wb_p[wb_p.sheetnames[0]]
    pdata = list(pws.iter_rows(values_only=True))
    meta = {
        'empresa': pdata[0][9] if pdata[0][9] else 2,
        'codVend': str(pdata[2][6]) if pdata[2][6] else '',
        'codCond': str(pdata[3][6]) if pdata[3][6] else '',
        'vendedor': _cel_para_texto(pdata[2][8]),
        'telefone': _cel_para_texto(pdata[2][9]),
        # Campos de cabeçalho (linhas 3-6, coluna C) — usados pelos clientes
        # 'Central' (sem tabela de filiais M:N:O): CNPJ/Filial/Endereço únicos,
        # preenchidos direto aqui em vez de vir de uma seleção de loja.
        'clienteNomePerfil': str(pdata[2][2]).strip() if pdata[2][2] else '',
        'cnpjPerfil': str(pdata[3][2]).strip() if pdata[3][2] else '',
        'filialPerfil': str(pdata[4][2]).strip() if pdata[4][2] else '',
        'enderecoPerfil': str(pdata[5][2]).strip() if pdata[5][2] else '',
    }
    produtos = []
    for r in pdata[7:]:
        if not r or not r[2]:
            break
        produtos.append({
            'empresa': int(r[0]) if r[0] in (1, 2) else None,  # coluna A: Fat.
            'codInterno': r[1],
            'nomePerfil': str(r[2]).strip(),
            'formato': str(r[3] or '').strip(),
            'embalagem': str(r[4]).strip(),
            'kgCx': float(r[6] or 20),
            'unidFat': str(r[7] or 'kg').strip(),
            'precoUnit': float(r[8] or 0),
            'obs': str(r[9] or '').strip(),
        })
    return meta, produtos



def _cel_para_texto(v):
    """Converte valor de célula do Excel pra texto de forma segura.

    FIX (23/08/2026): se a célula estiver formatada como número decimal
    (float) em vez de texto -- comum em telefone, já que alguém pode
    digitar sem formatar a célula como texto primeiro -- str(v) direto
    preserva o '.0' do float (ex.: 21973231111.0 -> '21973231111.0'),
    o que corrompe o telefone na hora de limpar caracteres não-numéricos
    (o '.0' vira um '0' grudado no final, um dígito a mais, número
    inválido). Detecta esse caso e converte pra int primeiro."""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _normaliza_cnpj(cnpj):
    """Remove pontuação do CNPJ para comparação confiável. Aceita tanto
    texto ('31.698.759/0001-13') quanto número puro (a célula do Excel pode
    vir como int quando não há formatação de texto aplicada)."""
    return ''.join(c for c in str(cnpj or '') if c.isdigit())


def _coord(v):
    """Lat/Lng do perfil, tolerante a como o usuário digitou no Excel: aceita
    número (-22.83) OU texto com vírgula decimal (-22,83, comum no BR). Valor
    vazio/ruim vira None (a filial fica sem pino no mapa em vez de derrubar a
    leitura inteira do perfil)."""
    if v is None:
        return None
    try:
        return float(str(v).strip().replace(',', '.'))
    except (ValueError, TypeError):
        return None


def ler_filiais(perfil_bytes):
    """Lê a tabela de filiais do Perfil Excel (colunas M:N:O, a partir da
    linha 9): CNPJ | Nome Filial | Número Filial. Opcionalmente também lê
    Endereço (col P), Cidade (col Q) e Região (col R), usadas pelo fluxo de
    pedido manual e pelo card de Região no PDF de expedição — para clientes
    que não têm essas colunas, ficam como string vazia.

    ATUALIZADO (22/08/2026): colunas U em diante = até 3 GRUPOS de 4
    colunas por encarregado da filial: nome, telefone, data de aniversário,
    time de futebol (U/V/W/X, Y/Z/AA/AB, AC/AD/AE/AF). Nome+telefone
    continuam obrigatórios pra o grupo contar (é o que o canal operacional
    de WhatsApp usa — whatsapp.py); aniversário e time são OPCIONAIS e só
    servem pra alimentar campos personalizados/segmentação de campanha
    dentro do próprio Botconversa (uso B da Planta) — nunca são usados em
    nenhuma lógica do nosso backend.

    Retorna dict {cnpj_normalizado: {'nome', 'numero', 'endereco', 'cidade',
    'regiao', 'lat', 'lng', 'encarregados': [{'nome','telefone',
    'dataAniversario','timeFutebol'}, ...]}}.
    Usado para enriquecer pedidos que só trazem CNPJ (Atacadão) ou
    CNPJ+nome (DOM) com o número de filial cadastrado uma única vez no perfil,
    e para alimentar o dropdown de filiais no fluxo manual."""
    wb_p = openpyxl.load_workbook(io.BytesIO(perfil_bytes), data_only=True)
    pws = wb_p[wb_p.sheetnames[0]]
    pdata = list(pws.iter_rows(values_only=True))
    filiais = {}
    for r in pdata[8:]:  # a partir da linha 9 (índice 8)
        if not r or len(r) < 15:
            continue
        cnpj_raw, nome, numero = r[12], r[13], r[14]  # colunas M, N, O
        endereco = r[15] if len(r) > 15 else None  # coluna P (opcional)
        cidade = r[16] if len(r) > 16 else None  # coluna Q (opcional)
        regiao = r[17] if len(r) > 17 else None  # coluna R = região (opcional)
        lat = r[18] if len(r) > 18 else None  # coluna S
        lng = r[19] if len(r) > 19 else None  # coluna T
        if not cnpj_raw:
            continue
        cnpj_norm = _normaliza_cnpj(cnpj_raw)
        if not cnpj_norm:
            continue

        # ATUALIZADO: até 3 grupos de 4 colunas por encarregado
        # (U/V/W/X, Y/Z/AA/AB, AC/AD/AE/AF = nome/telefone/aniversário/time)
        # FIX (23/08/2026): telefone usa _cel_para_texto (mesma proteção
        # do telefone do vendedor) -- célula numérica sem formatação de
        # texto corromperia o telefone (".0" de float virando dígito extra).
        encarregados = []
        for i_nome, i_tel, i_aniv, i_time in ((20, 21, 22, 23), (24, 25, 26, 27), (28, 29, 30, 31)):
            if len(r) <= i_time:
                continue
            enc_nome = _cel_para_texto(r[i_nome])
            enc_tel = _cel_para_texto(r[i_tel])
            enc_aniv = _cel_para_texto(r[i_aniv])
            enc_time = _cel_para_texto(r[i_time])
            if enc_nome and enc_tel:  # nome+telefone obrigatórios; aniversário/time opcionais
                encarregados.append({
                    'nome': enc_nome,
                    'telefone': enc_tel,
                    'dataAniversario': enc_aniv,
                    'timeFutebol': enc_time,
                })

        filiais[cnpj_norm] = {
            'nome': str(nome or '').strip(),
            'numero': numero,
            'endereco': str(endereco or '').strip(),
            'cidade': str(cidade or '').strip(),
            'regiao': str(regiao or '').strip(),
            'lat': _coord(lat),
            'lng': _coord(lng),
            'encarregados': encarregados,
        }
    return filiais


def buscar_filial(cnpj, filiais_map):
    """Busca nome e número de filial pelo CNPJ extraído do pedido.
    Retorna (nome, numero) ou (None, None) se não encontrado."""
    cnpj_norm = _normaliza_cnpj(cnpj)
    info = filiais_map.get(cnpj_norm)
    if info:
        return info['nome'], info['numero']
    return None, None


def ler_operadores(perfil_bytes):
    """Lê a lista de operadores (coluna L, a partir da linha 9) do Perfil.
    Usado nos clientes de pedido manual (sem PDF, ex: Guanabara Lojas), para
    alimentar o dropdown de quem está lançando o pedido no popup. Cresce
    livremente, sem depender do tamanho da tabela de produtos ou de filiais
    — cada lista (produtos, operadores, filiais) avança na sua própria
    coluna, independente das outras."""
    wb_p = openpyxl.load_workbook(io.BytesIO(perfil_bytes), data_only=True)
    pws = wb_p[wb_p.sheetnames[0]]
    pdata = list(pws.iter_rows(values_only=True))
    operadores = []
    for r in pdata[8:]:  # a partir da linha 9 (mesmo início da tabela de filiais)
        if not r or len(r) < 12:
            continue
        nome = r[11]  # coluna L
        if nome:
            operadores.append(str(nome).strip())
    return operadores


def processar_item(cod_cli, nome_raw, emb_tipo, qtde_emb, qtde_ped, preco, total, produtos):
    """Normaliza um item do pedido casando com o perfil e calculando
    kg planejados, número de caixas e demais campos derivados.

    FIX (01/08/2026): antes, quando match_perfil não achava o produto,
    o item era criado assim mesmo com dados "de emergência" (kgCx=20,
    embalagem chutada, nome cru do PDF) — aparecia como "(SEM MASTER)"
    no PDF/Excel, mas o pedido inteiro processava normal, silenciosamente.
    Combinado com o match aproximado antigo, isso também deixava passar
    itens colados no produto ERRADO sem nenhum aviso. Agora, sem
    correspondência exata no perfil, para o processamento do pedido
    inteiro com um erro claro — melhor corrigir o perfil (ou o nome do
    produto nele) do que gerar um romaneio errado ou incompleto."""
    nome_raw = re.sub(r'\s+', ' ', nome_raw).strip()
    pf = match_perfil(nome_raw, produtos)
    if pf is None:
        raise ValueError(f'Produto não cadastrado no perfil do mercado: "{nome_raw}"')
    kgCx = pf['kgCx']
    embalagem = pf['embalagem']
    if emb_tipo in ['CX', 'CXA']:
        kgPlan = qtde_ped * kgCx
        nrCx = qtde_ped
        qtdeMult = qtde_ped
        unidFat = 'cx'
    else:
        kgPlan = qtde_ped
        nrCx = round(kgPlan / kgCx, 1) if kgCx else 0
        qtdeMult = kgPlan
        unidFat = 'kg'
    return {
        'empresa': pf.get('empresa'),  # herda do perfil (coluna A)
        'codInterno': pf['codInterno'],
        'nomeProduto': pf['nomePerfil'],
        'formato': pf.get('formato', ''),
        'embalagem': embalagem,
        'kgCx': kgCx,
        'kgPlanejados': kgPlan,
        'nrCaixas': nrCx,
        'obs': pf.get('obs', ''),
        'qtdeMultipl': qtdeMult,
        'precoUnit': preco,
        'valorPedido': total,
        'precoSistema': pf['precoUnit'],
        'unidFat': unidFat,
    }
